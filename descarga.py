"""Generación del archivo Excel del costeo (botón de descarga).

Layout simple y legible (NO replica la plantilla original 1:1):
  1. Encabezado con los datos del embarque (orden, pedimento, TC, USD inputs).
  2. RESUMEN con los 4 números clave (total embarque, factura MXN, gastos, %).
  3. GASTOS DE IMPORTACIÓN: tabla con Sin IVA / IVA / Total por concepto.
  4. COSTO DEL EMBARQUE: cómo se arma el total a partir de factura + prorrateo.
  5. DETALLE DE EXTRACCIÓN: contribuciones del pedimento y tarifas elegidas.

Todas las cuentas son fórmulas (no valores hardcodeados): el archivo sigue
siendo editable y se recalcula al cambiar cualquier monto o el TC.
"""

from datetime import date
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Estilos
TITULO_FONT = Font(bold=True, size=16, color="FFFFFF")
TITULO_FILL = PatternFill("solid", start_color="2F5496")
SECCION_FONT = Font(bold=True, size=11, color="FFFFFF")
SECCION_FILL = PatternFill("solid", start_color="4472C4")
HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill("solid", start_color="DDEBF7")
ETIQUETA_FONT = Font(bold=True, size=10)
TOTAL_FONT = Font(bold=True, size=12)
RESALTAR_FONT = Font(bold=True, size=14, color="1F4E78")
RESALTAR_FILL = PatternFill("solid", start_color="E2EFDA")
NOTA_FONT = Font(italic=True, size=9, color="595959")

CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

FORMATO_MXN = "$#,##0.00"
FORMATO_PCT = "0.00%"
FORMATO_TC = "0.0000"


def _seccion(ws, fila, texto, columnas=4):
    """Pinta un header de sección (merged, color)."""
    last_col = get_column_letter(columnas)
    ws.merge_cells(f"A{fila}:{last_col}{fila}")
    cell = ws[f"A{fila}"]
    cell.value = texto
    cell.font = SECCION_FONT
    cell.fill = SECCION_FILL
    cell.alignment = CENTER


def generar_xlsx(inp, r, detalle_pedimento=None, mh_label=None, ft_label=None, ingresos=None):
    """Construye el .xlsx del costeo y devuelve sus bytes.

    Args:
        inp: CosteoInputs con los valores capturados.
        r: CosteoResultado (los totales se recalculan con fórmulas).
        detalle_pedimento: dict opcional con dta/prv/igi/iva/iva_prv/ieps...
        mh_label: etiqueta del selector de maniobras (ej. "WISE ($20,640.00)").
        ft_label: etiqueta del selector de flete terrestre.
        ingresos: lista opcional de folios de ingreso 'Done' (ej. MV1/IN/13806).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COSTEO"

    # --- Título ---
    ws.merge_cells("A1:D1")
    ws["A1"] = "COSTEO DE IMPORTACIÓN"
    ws["A1"].font = TITULO_FONT
    ws["A1"].fill = TITULO_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    # --- Datos del embarque (fila 3-10). Los formularios referencian a estas celdas. ---
    # OJO: el orden de estas filas es importante. Las fórmulas usan referencias
    # fijas: B6 = tipo de cambio, B7 = factura USD, B8 = flete USD, B9 = cargos.
    # Si agregas filas, hazlo DESPUÉS de la fila 9 para no romper esas referencias.
    datos = [
        ("Órdenes", inp.orden or "", None),
        ("No. pedimento", inp.no_pedimento or "", None),
        ("Proveedor", inp.proveedor or "", None),
        ("Tipo de cambio", float(inp.tipo_cambio or 0), FORMATO_TC),
        ("Factura proveedor (USD)", float(inp.factura_proveedor_usd or 0), FORMATO_MXN),
        ("Flete marítimo (USD)", float(inp.flete_maritimo_usd or 0), FORMATO_MXN),
        ("Cargos por transferencia (USD)", float(inp.cargos_transferencia_usd or 0), FORMATO_MXN),
        ("Fecha", date.today(), "yyyy-mm-dd"),
    ]
    for i, (etiqueta, valor, fmt) in enumerate(datos, start=3):
        ws[f"A{i}"] = etiqueta
        ws[f"A{i}"].font = ETIQUETA_FONT
        ws[f"B{i}"] = valor
        if fmt:
            ws[f"B{i}"].number_format = fmt
        ws.merge_cells(f"B{i}:D{i}")
    # Referencias clave (para uso interno en fórmulas):
    #   B6 = tipo de cambio
    #   B7 = factura proveedor (USD)
    #   B8 = flete marítimo (USD)
    #   B9 = cargos por transferencia (USD)

    # --- RESUMEN ---
    _seccion(ws, 12, "RESUMEN")
    resumen = [
        ("Total de embarque (MXN)", "=B34", FORMATO_MXN, True),
        ("Factura del proveedor (MXN)", "=B7*B6", FORMATO_MXN, False),
        ("Gastos de importación (MXN, sin IVA)", "=B27", FORMATO_MXN, False),
        ("% gasto vs factura", "=IFERROR(B15/B14,0)", FORMATO_PCT, False),
    ]
    for i, (etiqueta, formula, fmt, resaltar) in enumerate(resumen, start=13):
        ws[f"A{i}"] = etiqueta
        ws[f"A{i}"].font = ETIQUETA_FONT
        ws[f"B{i}"] = formula
        ws[f"B{i}"].number_format = fmt
        if resaltar:
            ws[f"B{i}"].font = RESALTAR_FONT
            ws[f"B{i}"].fill = RESALTAR_FILL
            ws[f"A{i}"].fill = RESALTAR_FILL
            ws.row_dimensions[i].height = 24
        ws.merge_cells(f"B{i}:D{i}")

    # --- GASTOS DE IMPORTACIÓN ---
    _seccion(ws, 18, "GASTOS DE IMPORTACIÓN")
    for col_idx, label in enumerate(["Concepto", "Sin IVA (MXN)", "IVA (MXN)", "Total (MXN)"], start=1):
        cell = ws.cell(row=19, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    # (label, sin_iva, iva)  -- iva=None significa "calcular como sin_iva * 16%"
    gastos = [
        ("Impuestos del pedimento", float(inp.impuestos or 0), float(inp.iva_aduana or 0)),
        ("Maniobras y honorarios", float(inp.maniobras or 0), None),
        ("Demoras", float(inp.otros_demoras or 0), None),
        ("Almacenaje", float(inp.almacenajes or 0), None),
        ("Flete marítimo", "=B8*B6", 0),
        ("Flete local", float(inp.flete_local or 0), None),
        ("Pos. en falso", float(inp.falsos or 0), 0),
    ]
    for i, (label, sin_iva, iva) in enumerate(gastos, start=20):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = sin_iva
        ws[f"C{i}"] = f"=B{i}*0.16" if iva is None else iva
        ws[f"D{i}"] = f"=B{i}+C{i}"
        for col in "BCD":
            ws[f"{col}{i}"].number_format = FORMATO_MXN

    # Total de gastos
    ws["A27"] = "TOTAL GASTOS"
    ws["A27"].font = TOTAL_FONT
    ws["A27"].fill = HEADER_FILL
    ws["B27"] = "=SUM(B20:B26)"
    ws["C27"] = "=SUM(C20:C26)"
    ws["D27"] = "=B27+C27"
    for col in "BCD":
        ws[f"{col}27"].font = TOTAL_FONT
        ws[f"{col}27"].fill = HEADER_FILL
        ws[f"{col}27"].number_format = FORMATO_MXN

    # --- COSTO DEL EMBARQUE ---
    _seccion(ws, 29, "COSTO DEL EMBARQUE")
    ws["A30"] = "Concepto"
    ws["B30"] = "Monto (MXN)"
    ws["A30"].font = HEADER_FONT
    ws["B30"].font = HEADER_FONT
    ws["A30"].fill = HEADER_FILL
    ws["B30"].fill = HEADER_FILL
    ws["B30"].alignment = CENTER
    ws.merge_cells("B30:D30")

    embarque = [
        ("Factura del proveedor", "=B7*B6"),
        ("Cargos por transferencia", "=B9*B6"),
        ("Prorrateo de gastos (sin IVA)", "=B27"),
    ]
    for i, (label, formula) in enumerate(embarque, start=31):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = formula
        ws[f"B{i}"].number_format = FORMATO_MXN
        ws.merge_cells(f"B{i}:D{i}")

    ws["A34"] = "TOTAL DE EMBARQUE"
    ws["A34"].font = RESALTAR_FONT
    ws["A34"].fill = RESALTAR_FILL
    ws["B34"] = "=SUM(B31:B33)"
    ws["B34"].font = RESALTAR_FONT
    ws["B34"].fill = RESALTAR_FILL
    ws["B34"].number_format = FORMATO_MXN
    ws.row_dimensions[34].height = 24
    ws.merge_cells("B34:D34")

    # --- Nota ---
    ws.merge_cells("A36:D36")
    ws["A36"] = ("Nota: el IVA es recuperable y NO se carga al costo. "
                 "Solo los gastos netos (sin IVA) se prorratean sobre la mercancía.")
    ws["A36"].font = NOTA_FONT
    ws["A36"].alignment = CENTER_WRAP
    ws.row_dimensions[36].height = 30

    fila = 38

    # --- Landed cost (sección propia; vacía si no se capturó) ---
    _seccion(ws, fila, "LANDED COST")
    fila += 1
    ws[f"A{fila}"] = inp.landed_cost or ""
    ws.merge_cells(f"A{fila}:D{fila}")
    fila += 2

    # --- Ingresos (Done) relacionados a las órdenes ---
    if ingresos:
        _seccion(ws, fila, "INGRESOS (Done)")
        fila += 1
        for folio in ingresos:
            ws[f"A{fila}"] = folio
            ws.merge_cells(f"A{fila}:D{fila}")
            fila += 1
        fila += 1

    # --- Detalle de extracción ---
    _seccion(ws, fila, "DETALLE DE EXTRACCIÓN")
    fila += 1

    if detalle_pedimento and any(detalle_pedimento.get(k) for k in
                                  ("dta", "prv", "igi", "ieps", "iva", "iva_prv", "ieps_iva")):
        ws[f"A{fila}"] = "Contribuciones extraídas del pedimento:"
        ws[f"A{fila}"].font = ETIQUETA_FONT
        ws.merge_cells(f"A{fila}:D{fila}")
        fila += 1
        for label, key in [("DTA", "dta"), ("PRV", "prv"), ("IGI", "igi"),
                           ("IEPS", "ieps"), ("IVA", "iva"),
                           ("IVA/PRV", "iva_prv"), ("IEPS/IVA", "ieps_iva")]:
            val = detalle_pedimento.get(key, 0) or 0
            if val:
                ws[f"A{fila}"] = label
                ws[f"B{fila}"] = float(val)
                ws[f"B{fila}"].number_format = FORMATO_MXN
                ws.merge_cells(f"B{fila}:D{fila}")
                fila += 1
        fila += 1

    if mh_label or ft_label:
        ws[f"A{fila}"] = "Tarifas elegidas del catálogo:"
        ws[f"A{fila}"].font = ETIQUETA_FONT
        ws.merge_cells(f"A{fila}:D{fila}")
        fila += 1
        if mh_label:
            ws[f"A{fila}"] = "Maniobras y honorarios"
            ws[f"B{fila}"] = mh_label
            ws.merge_cells(f"B{fila}:D{fila}")
            fila += 1
        if ft_label:
            ws[f"A{fila}"] = "Flete local"
            ws[f"B{fila}"] = ft_label
            ws.merge_cells(f"B{fila}:D{fila}")
            fila += 1

    # --- Anchos de columna ---
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
